from openpyxl import load_workbook
import time
import xlwt
# file = 'table.xlsx'

arr_EN = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

class Excel():
    def get_file(self, file):
        # file = "./schedule-mi-3.xlsx"
        try:
            wb = load_workbook(file)
            self.wb = wb
            return wb
        except Exception as ex:
            print("Something went wrong while loading the file.")
            time.sleep(3)
            raise SystemExit(1)

    def get_sheet(self, sheet):

            # for i in self.wb.sheetnames:
            #     print(f"{i}", end=" ")
            # print()

            sheetName = sheet
            try:
                sheet = self.wb[sheetName]
                self.sheet = sheet
                return sheet

            except Exception as ex:
                print("Something went wrong with a sheet.")
                time.sleep(3)
                raise SystemExit(1)

    def get_data(self, field, row):
        return self.sheet[f"{field}{row}"].value
    def get_data_by_text(self, a):
        return self.sheet[a].value
    
    def get_formation_sheet(self, formation):
        sheets = {
            'L1-MI1': 'MI-1',
            'L1-MI2': 'MI-2',
            'L1-MI3': 'MI-3',
            'L1-MI4': 'MI-4',

            'LDD1-IM1': 'LDD1 IM-1',
            'LDD1-IM2': 'LDD1 IM-2',

            'L1-MP-A1': 'A1',
            'L1-MP-A2': 'A2',
            'L1-MP-A3': 'A3',
            'L1-MP-B1': 'B1',
            'L1-MP-B2': 'B2',

            'LDD1-MPSI-B3': 'MPSI B3',
            'LDD1-MPSI-B4': 'MPSI B4',
            'LDD1-MPSI-STAPS-B5': 'LDD1 STAPS-SPI',
            'LDD2-MPSI-STAPS': 'LDD2 STAPS-SPI',
        }
        return sheets[formation]
    
    def get_formation_file(self, formation):
        files = {
            'L1-MI1': './Semainier_MI_IM_S2_23-24.xlsx',
            'L1-MI2': './Semainier_MI_IM_S2_23-24.xlsx',
            'L1-MI3': './Semainier_MI_IM_S2_23-24.xlsx',
            'L1-MI4': './Semainier_MI_IM_S2_23-24.xlsx',

            'LDD1-IM1': './Semainier_MI_IM_S2_23-24.xlsx',
            'LDD1-IM2': './Semainier_MI_IM_S2_23-24.xlsx',

            'L1-MP-A1': './Semainier_MP_MPSI_STAPS S2_23-24.xlsx',
            'L1-MP-A2': './Semainier_MP_MPSI_STAPS S2_23-24.xlsx',
            'L1-MP-A3': './Semainier_MP_MPSI_STAPS S2_23-24.xlsx',
            'L1-MP-B1': './Semainier_MP_MPSI_STAPS S2_23-24.xlsx',
            'L1-MP-B2': './Semainier_MP_MPSI_STAPS S2_23-24.xlsx',

            'LDD1-MPSI-B3': './Semainier_MP_MPSI_STAPS S2_23-24.xlsx',
            'LDD1-MPSI-B4': './Semainier_MP_MPSI_STAPS S2_23-24.xlsx',
            'LDD1-MPSI-STAPS-B5': './Semainier_MP_MPSI_STAPS S2_23-24.xlsx',
            'LDD2-MPSI-STAPS': './Semainier_MP_MPSI_STAPS S2_23-24.xlsx',
        }
        return files[formation]